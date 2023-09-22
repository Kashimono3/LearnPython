from openpyxl import *
from tkinter import *
from tkinter import ttk
wb = load_workbook(r'C:\Users\cict\Desktop\Python\dssv.xlsx')


sheet = wb.active


def excel():
	

	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50
	sheet.column_dimensions['H'].width = 50


	sheet.cell(row=1, column=1).value = "Mã số sinh viên"
	sheet.cell(row=1, column=2).value = "Họ tên"
	sheet.cell(row=1, column=3).value = "Ngày sinh"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "Số điện thoại"
	sheet.cell(row=1, column=6).value = "Học kỳ"
	sheet.cell(row=1, column=7).value = "Năm học"
	sheet.cell(row=1, column=8).value = "Môn học"


def focus1(event):

	mssv.focus_set()



def focus2(event):
	
	hoTen.focus_set()



def focus3(event):

	ngaySinh.focus_set()



def focus4(event):

	email.focus_set()



def focus5(event):

	soDienThoai.focus_set()



def focus6(event):

	hocKy.focus_set()
def focus7(event):

	namHoc.focus_set()
def focus8(event):

	monHoc.focus_set()


def insert():
	

	if (mssv.get() == "" and
		hoTen.get() == "" and
		ngaySinh.get() == "" and
		email.get() == "" and
		soDienThoai.get() == "" and
		hocKy.get() == "" and
		namHoc.get() == "" and
		monHoc.get() == "" ):
			
		print("empty input")

	else:

		current_row = sheet.max_row
		current_column = sheet.max_column


		sheet.cell(row=current_row + 1, column=1).value = mssv.get()
		sheet.cell(row=current_row + 1, column=2).value = hoTen.get()
		sheet.cell(row=current_row + 1, column=3).value = ngaySinh.get()
		sheet.cell(row=current_row + 1, column=4).value = email.get()
		sheet.cell(row=current_row + 1, column=5).value = soDienThoai.get()
		sheet.cell(row=current_row + 1, column=6).value = hocKy.get()
		sheet.cell(row=current_row + 1, column=7).value = namHoc.get()
		sheet.cell(row=current_row + 1, column=8).value = monHoc.get()

		wb.save(r'C:\Users\cict\Desktop\Python\dssv.xlsx')


		mssv.focus_set()

	
		clear()



if __name__ == "__main__":
	

	root = Tk()

	root.configure(background='light green')


	root.title("registration form")

	
	root.geometry("500x300")

	excel()


	heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", fg="red",bg="light green",font ="Roboto 15 bold")

	mssv = Label(root, text="Mã số sinh viên", bg="light green")


	hoTen = Label(root, text="Họ tên", bg="light green")


	ngaySinh = Label(root, text="Ngày sinh", bg="light green")


	email = Label(root, text="Email", bg="light green")


	soDienThoai = Label(root, text="Số điện thoại", bg="light green")


	hocKy = Label(root, text="Học Kỳ", bg="light green")
	namHoc = Label(root,text = "Năm học", bg= "light green")

	n = StringVar()
	namHoccbb = ttk.Combobox(root, textvariable = n, width  = 27)
	namHoccbb ["values"]=("2020","2021","2022")


	heading.grid(row=0, column=1)
	mssv.grid(row=1, column=0)
	hoTen.grid(row=2, column=0)
	ngaySinh.grid(row=3, column=0)
	email.grid(row=4, column=0)
	soDienThoai.grid(row=5, column=0)
	hocKy.grid(row=6, column=0)
	namHoc.grid(row=10, column=0)
	namHoccbb.grid(row = 10, column = 1)
	namHoccbb.current()

	mssv = Entry(root)
	hoTen = Entry(root)
	ngaySinh = Entry(root)
	email = Entry(root)
	soDienThoai = Entry(root)
	hocKy = Entry(root)
	namHoc = Entry(root)
	monHoc = Entry(root)



	mssv.bind("<Return>", focus1)


	hoTen.bind("<Return>", focus2)


	ngaySinh.bind("<Return>", focus3)


	email.bind("<Return>", focus4)


	soDienThoai.bind("<Return>", focus5)


	hocKy.bind("<Return>", focus6)

	namHoc.bind("<Return>", focus7)

	
	monHoc.bind("<Return>", focus8)

	mssv.grid(row=1, column=1, ipadx="100")
	hoTen.grid(row=2, column=1, ipadx="100")
	ngaySinh.grid(row=3, column=1, ipadx="100")
	email.grid(row=4, column=1, ipadx="100")
	soDienThoai.grid(row=5, column=1, ipadx="100")
	hocKy.grid(row=6, column=1, ipadx="100")
	
	monHoc.grid(row=8, column=1, ipadx="100")


	excel()


	submit = Button(root, text="Đăng Ký", fg="Black",
							bg="green", command=insert)
	submit.grid(row=9, column=1)


	root.mainloop()
